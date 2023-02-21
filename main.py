import openpyxl

# define myfunc() function
def myfunc(a, b):
    return a + b

def myfunc(input_file_path):
    wb = openpyxl.load_workbook(input_file_path)
    sheet = wb.active
    # Prompt the user to enter the input value from the Excel file
    user_input = input(f"Enter any input from the Excel file (or press Enter to use the default value): ")
    if user_input:
        input_value = sheet[user_input].value
    else:
        input_value = sheet['A1'].value  # Use default value if user didn't enter a value
    # Do some processing with the input value
    result = input_value * 2
    return result
    
#def read_input_from_excel():
    # open the Excel file
    #wb = openpyxl.load_workbook('./input_data.xlsx')

    # select the active sheet
    #ws = wb.active

    # prompt the user to enter input
    try:
        user_input = input("Enter any input from the Excel file (or press Enter to use the default value): ") or "default value"
    except EOFError:
        user_input = "default value"

    # search for the input in the sheet
    for row in ws.iter_rows(values_only=True):
        if user_input in row:
            return row
    return None

